#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
发票 OCR 识别工具 - GUI 图形界面版
功能：通过图形界面选择 PDF 文件，自动识别并生成 Excel
作者：OpenClaw
版本：1.0
日期：2026-04-23
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
import sys

# 依赖检查
try:
    import fitz
    import pytesseract
    from PIL import Image
    import io
    import re
    import json
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError as e:
    print(f"错误：缺少依赖库 - {e}")
    print("请运行：pip3 install pymupdf pytesseract pillow openpyxl")
    sys.exit(1)

class InvoiceOCRApp:
    """发票 OCR 图形界面应用"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("发票 OCR 识别工具 v1.0")
        self.root.geometry("800x600")
        self.root.minsize(700, 500)
        
        # 文件列表
        self.pdf_files = []
        self.output_dir = tk.StringVar()
        self.output_dir.set(os.path.expanduser("~"))
        
        # 创建界面
        self.create_widgets()
    
    def create_widgets(self):
        """创建界面组件"""
        # 标题
        title_frame = ttk.Frame(self.root)
        title_frame.pack(fill=tk.X, padx=10, pady=10)
        
        title_label = ttk.Label(
            title_frame,
            text="📄 发票 OCR 识别工具",
            font=("Arial", 18, "bold")
        )
        title_label.pack(side=tk.LEFT)
        
        subtitle_label = ttk.Label(
            title_frame,
            text="v1.0 - 自动识别发票信息并生成 Excel",
            font=("Arial", 10)
        )
        subtitle_label.pack(side=tk.LEFT, padx=10)
        
        # 文件选择区域
        file_frame = ttk.LabelFrame(self.root, text="PDF 文件选择", padding=10)
        file_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # 按钮区域
        btn_frame = ttk.Frame(file_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(
            btn_frame,
            text="➕ 添加 PDF 文件",
            command=self.add_files
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            btn_frame,
            text="📁 添加文件夹",
            command=self.add_folder
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            btn_frame,
            text="🗑️ 清空列表",
            command=self.clear_files
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            btn_frame,
            text="📂 输出目录:",
            command=self.select_output_dir
        ).pack(side=tk.LEFT, padx=(20, 5))
        
        ttk.Label(
            btn_frame,
            textvariable=self.output_dir,
            wraplength=300
        ).pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # 文件列表
        list_frame = ttk.Frame(file_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        # 创建 Treeview
        columns = ("序号", "文件名", "大小", "状态")
        self.file_tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=10)
        
        for col in columns:
            self.file_tree.heading(col, text=col)
            self.file_tree.column(col, width=100 if col != "文件名" else 400)
        
        # 滚动条
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.file_tree.yview)
        self.file_tree.configure(yscrollcommand=scrollbar.set)
        
        self.file_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 进度区域
        progress_frame = ttk.LabelFrame(self.root, text="处理进度", padding=10)
        progress_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.progress_bar = ttk.Progressbar(progress_frame, mode='indeterminate', length=600)
        self.progress_bar.pack(fill=tk.X, pady=(0, 10))
        
        self.progress_label = ttk.Label(progress_frame, text="就绪", font=("Arial", 10))
        self.progress_label.pack(anchor=tk.W)
        
        # 日志区域
        log_frame = ttk.LabelFrame(self.root, text="处理日志", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=8, font=("Consolas", 9))
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # 底部按钮
        bottom_frame = ttk.Frame(self.root)
        bottom_frame.pack(fill=tk.X, padx=10, pady=10)
        
        self.start_btn = ttk.Button(
            bottom_frame,
            text="🚀 开始识别",
            command=self.start_processing,
            style="Accent.TButton"
        )
        self.start_btn.pack(side=tk.RIGHT, padx=5)
        
        # 统计信息
        self.stats_label = ttk.Label(
            bottom_frame,
            text="文件数：0 | 总大小：0 MB",
            font=("Arial", 10)
        )
        self.stats_label.pack(side=tk.LEFT)
    
    def log(self, message):
        """添加日志"""
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def add_files(self):
        """添加 PDF 文件"""
        files = filedialog.askopenfilenames(
            title="选择 PDF 文件",
            filetypes=[("PDF 文件", "*.pdf"), ("所有文件", "*.*")]
        )
        
        for file in files:
            if file not in self.pdf_files:
                size = os.path.getsize(file) / 1024 / 1024  # MB
                self.pdf_files.append(file)
                idx = len(self.pdf_files)
                self.file_tree.insert("", tk.END, values=(
                    idx,
                    os.path.basename(file),
                    f"{size:.2f} MB",
                    "待处理"
                ))
        
        self.update_stats()
    
    def add_folder(self):
        """添加文件夹中的所有 PDF"""
        folder = filedialog.askdirectory(title="选择包含 PDF 的文件夹")
        if not folder:
            return
        
        pdf_count = 0
        for file in os.listdir(folder):
            if file.lower().endswith('.pdf'):
                full_path = os.path.join(folder, file)
                if full_path not in self.pdf_files:
                    size = os.path.getsize(full_path) / 1024 / 1024
                    self.pdf_files.append(full_path)
                    idx = len(self.pdf_files)
                    self.file_tree.insert("", tk.END, values=(
                        idx,
                        file,
                        f"{size:.2f} MB",
                        "待处理"
                    ))
                    pdf_count += 1
        
        self.log(f"从文件夹添加了 {pdf_count} 个 PDF 文件")
        self.update_stats()
    
    def clear_files(self):
        """清空文件列表"""
        self.pdf_files = []
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
        self.update_stats()
        self.log("已清空文件列表")
    
    def select_output_dir(self):
        """选择输出目录"""
        directory = filedialog.askdirectory(title="选择输出目录")
        if directory:
            self.output_dir.set(directory)
            self.log(f"输出目录：{directory}")
    
    def update_stats(self):
        """更新统计信息"""
        count = len(self.pdf_files)
        total_size = sum(os.path.getsize(f) for f in self.pdf_files) / 1024 / 1024
        self.stats_label.config(text=f"文件数：{count} | 总大小：{total_size:.2f} MB")
    
    def start_processing(self):
        """开始处理"""
        if not self.pdf_files:
            messagebox.showwarning("提示", "请先添加 PDF 文件")
            return
        
        # 禁用按钮
        self.start_btn.config(state=tk.DISABLED)
        self.progress_bar.start()
        
        # 在新线程中处理
        thread = threading.Thread(target=self.process_files)
        thread.daemon = True
        thread.start()
    
    def process_files(self):
        """处理文件（在后台线程）"""
        try:
            total_files = len(self.pdf_files)
            self.log(f"\n开始处理 {total_files} 个文件...")
            
            for i, pdf_path in enumerate(self.pdf_files, 1):
                self.root.after(0, lambda p=pdf_path: self.log(f"\n[{i}/{total_files}] 处理：{os.path.basename(p)}"))
                self.root.after(0, lambda: self.file_tree.item(self.file_tree.get_children()[i-1], values=(
                    i,
                    os.path.basename(self.pdf_files[i-1]),
                    f"{os.path.getsize(self.pdf_files[i-1])/1024/1024:.2f} MB",
                    "处理中..."
                )))
                
                try:
                    result = self.process_single_pdf(pdf_path)
                    status = "✅ 完成" if result else "❌ 失败"
                except Exception as e:
                    self.root.after(0, lambda err=str(e): self.log(f"错误：{err}"))
                    status = f"❌ 错误"
                
                self.root.after(0, lambda idx=i-1, s=status: self.file_tree.item(
                    self.file_tree.get_children()[idx],
                    values=(
                        idx + 1,
                        os.path.basename(self.pdf_files[idx]),
                        f"{os.path.getsize(self.pdf_files[idx])/1024/1024:.2f} MB",
                        s
                    )
                ))
            
            self.root.after(0, lambda: self.log("\n✅ 所有文件处理完成！"))
            self.root.after(0, lambda: messagebox.showinfo("完成", f"成功处理 {total_files} 个文件！"))
        
        finally:
            self.root.after(0, self.progress_bar.stop)
            self.root.after(0, lambda: self.start_btn.config(state=tk.NORMAL))
    
    def process_single_pdf(self, pdf_path):
        """处理单个 PDF"""
        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
        output_dir = self.output_dir.get()
        
        # 打开 PDF
        doc = fitz.open(pdf_path)
        total_pages = len(doc)
        self.root.after(0, lambda: self.log(f"  总页数：{total_pages}"))
        
        # 处理每一页
        invoices = []
        for page_idx, page in enumerate(doc):
            try:
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                img = Image.open(io.BytesIO(pix.tobytes("png")))
                text = pytesseract.image_to_string(img, lang='chi_sim+eng', config='--psm 6')
                info = self.extract_invoice_info(text)
                
                invoices.append({
                    "page": page_idx + 1,
                    **info
                })
                
                if (page_idx + 1) % 10 == 0:
                    self.root.after(0, lambda p=page_idx+1: self.log(f"  进度：{p}/{total_pages} 页"))
            except Exception as e:
                self.root.after(0, lambda err=str(e), p=page_idx+1: self.log(f"  警告：第{p}页处理失败 - {err}"))
                invoices.append({"page": page_idx + 1, "invoice_code": "", "invoice_no": "", "invoice_date": "", "amount": ""})
        
        doc.close()
        
        # 统计
        total = len(invoices)
        with_no = sum(1 for i in invoices if i.get("invoice_no"))
        with_date = sum(1 for i in invoices if i.get("invoice_date"))
        with_amount = sum(1 for i in invoices if i.get("amount"))
        
        amounts = []
        for i in invoices:
            amt = i.get("amount", "")
            if amt:
                try:
                    amounts.append(float(amt.replace(",", "")))
                except:
                    pass
        
        self.root.after(0, lambda: self.log(f"  识别结果:"))
        self.root.after(0, lambda: self.log(f"    发票号码：{with_no}/{total} ({with_no/total*100:.1f}%)"))
        self.root.after(0, lambda: self.log(f"    开票日期：{with_date}/{total} ({with_date/total*100:.1f}%)"))
        self.root.after(0, lambda: self.log(f"    金   额：{with_amount}/{total} ({with_amount/total*100:.1f}%)"))
        if amounts:
            self.root.after(0, lambda: self.log(f"    总金额：¥{sum(amounts):,.2f}"))
        
        # 保存 JSON
        json_path = os.path.join(output_dir, f"{base_name}_invoices.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(invoices, f, ensure_ascii=False, indent=2)
        self.root.after(0, lambda: self.log(f"  JSON 已保存：{os.path.basename(json_path)}"))
        
        # 生成 Excel
        excel_path = self.create_excel(invoices, output_dir, base_name)
        self.root.after(0, lambda: self.log(f"  Excel 已保存：{os.path.basename(excel_path)}"))
        
        return True
    
    def extract_invoice_info(self, text):
        """提取发票信息"""
        result = {
            "invoice_code": "",
            "invoice_no": "",
            "invoice_date": "",
            "amount": ""
        }
        
        match = re.search(r'No\s*(\d{8})', text)
        if match:
            result["invoice_no"] = match.group(1)
        
        match = re.search(r'(\d{12})', text)
        if match:
            result["invoice_code"] = match.group(1)
        
        match = re.search(r'(\d{4}年\d{2}月\d{2}日)', text)
        if match:
            result["invoice_date"] = match.group(1)
        
        match = re.search(r'[（(]小写 [)）]\s*¥?\s*([\d,]+\.?\d*)', text)
        if match:
            result["amount"] = match.group(1).replace(",", "")
        
        return result
    
    def create_excel(self, invoices, output_dir, base_name):
        """创建 Excel 表格"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "发票信息"
        
        headers = ["序号", "页码", "发票代码", "发票号码", "开票日期", "金额（元）", "备注"]
        ws.append(headers)
        
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row_num = 1
        for idx, inv in enumerate(invoices, 1):
            row_num += 1
            row = [
                idx,
                inv.get("page", ""),
                inv.get("invoice_code", ""),
                inv.get("invoice_no", ""),
                inv.get("invoice_date", ""),
                inv.get("amount", ""),
                "OCR 识别" if inv.get("invoice_no") else "部分信息缺失"
            ]
            ws.append(row)
            for cell in ws[row_num]:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 6
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.freeze_panes = 'A2'
        
        excel_path = os.path.join(output_dir, f"{base_name}_发票信息汇总表.xlsx")
        wb.save(excel_path)
        return excel_path


def main():
    """主函数"""
    root = tk.Tk()
    app = InvoiceOCRApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
