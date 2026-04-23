#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
发票 OCR 识别工具
功能：批量识别 PDF 凭证中的发票信息，生成 Excel 汇总表
作者：OpenClaw
版本：1.0
日期：2026-04-23
"""

import fitz
import pytesseract
from PIL import Image
import io
import re
import json
import sys
import os
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("警告：openpyxl 未安装，无法生成 Excel")

# 配置
BATCH_SIZE = 8  # 每批处理页数
PDF_DPI = 2  # PDF 转图像放大倍数

def extract_invoice_info(text):
    """从 OCR 文本中提取发票信息"""
    result = {
        "invoice_code": "",
        "invoice_no": "",
        "invoice_date": "",
        "amount": ""
    }
    
    # 发票号码
    match = re.search(r'No\s*(\d{8})', text)
    if match:
        result["invoice_no"] = match.group(1)
    
    # 发票代码
    match = re.search(r'(\d{12})', text)
    if match:
        result["invoice_code"] = match.group(1)
    
    # 开票日期
    match = re.search(r'(\d{4}年\d{2}月\d{2}日)', text)
    if match:
        result["invoice_date"] = match.group(1)
    
    # 金额
    match = re.search(r'[（(]小写 [)）]\s*¥?\s*([\d,]+\.?\d*)', text)
    if match:
        result["amount"] = match.group(1).replace(",", "")
    
    return result

def process_pdf(pdf_path, output_dir=None):
    """处理单个 PDF 文件"""
    print(f"\n处理文件：{pdf_path}")
    
    if not os.path.exists(pdf_path):
        print(f"错误：文件不存在 - {pdf_path}")
        return None
    
    # 输出目录
    if output_dir is None:
        output_dir = os.path.dirname(pdf_path)
    
    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    
    # 打开 PDF
    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        print(f"错误：无法打开 PDF - {e}")
        return None
    
    print(f"  总页数：{len(doc)}")
    
    # 处理每一页
    invoices = []
    for i, page in enumerate(doc):
        try:
            pix = page.get_pixmap(matrix=fitz.Matrix(PDF_DPI, PDF_DPI))
            img = Image.open(io.BytesIO(pix.tobytes("png")))
            text = pytesseract.image_to_string(img, lang='chi_sim+eng', config='--psm 6')
            info = extract_invoice_info(text)
            
            invoices.append({
                "page": i + 1,
                **info
            })
            
            if (i + 1) % BATCH_SIZE == 0:
                print(f"  进度：{i + 1}/{len(doc)} 页")
        except Exception as e:
            print(f"  警告：第{i+1}页处理失败 - {e}")
            invoices.append({"page": i + 1, "invoice_code": "", "invoice_no": "", "invoice_date": "", "amount": ""})
    
    doc.close()
    
    # 统计
    total = len(invoices)
    with_no = sum(1 for i in invoices if i.get("invoice_no"))
    with_date = sum(1 for i in invoices if i.get("invoice_date"))
    with_amount = sum(1 for i in invoices if i.get("amount"))
    
    amounts = []
    for i in invoices:
        amt = i.get("amount", "")
        if amt:
            try:
                amounts.append(float(amt.replace(",", "")))
            except:
                pass
    
    print(f"\n识别结果:")
    print(f"  总发票数：{total} 张")
    print(f"  发票号码：{with_no} 张 ({with_no/total*100:.1f}%)")
    print(f"  开票日期：{with_date} 张 ({with_date/total*100:.1f}%)")
    print(f"  金   额：{with_amount} 张 ({with_amount/total*100:.1f}%)")
    if amounts:
        print(f"  总金额：¥{sum(amounts):,.2f}")
    
    # 保存 JSON
    json_path = os.path.join(output_dir, f"{base_name}_invoices.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(invoices, f, ensure_ascii=False, indent=2)
    print(f"\nJSON 已保存：{json_path}")
    
    # 生成 Excel
    if HAS_OPENPYXL:
        excel_path = create_excel(invoices, output_dir, base_name)
        print(f"Excel 已保存：{excel_path}")
        return {"json": json_path, "excel": excel_path, "stats": {
            "total": total,
            "with_no": with_no,
            "with_date": with_date,
            "with_amount": with_amount,
            "total_amount": sum(amounts) if amounts else 0
        }}
    else:
        return {"json": json_path, "excel": None, "stats": {
            "total": total,
            "with_no": with_no,
            "with_date": with_date,
            "with_amount": with_amount,
            "total_amount": sum(amounts) if amounts else 0
        }}

def create_excel(invoices, output_dir, base_name):
    """创建 Excel 表格"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发票信息"
    
    # 表头
    headers = ["序号", "页码", "发票代码", "发票号码", "开票日期", "金额（元）", "备注"]
    ws.append(headers)
    
    # 样式
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 数据
    row_num = 1
    for idx, inv in enumerate(invoices, 1):
        row_num += 1
        row = [
            idx,
            inv.get("page", ""),
            inv.get("invoice_code", ""),
            inv.get("invoice_no", ""),
            inv.get("invoice_date", ""),
            inv.get("amount", ""),
            "OCR 识别" if inv.get("invoice_no") else "部分信息缺失"
        ]
        ws.append(row)
        for cell in ws[row_num]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # 列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.freeze_panes = 'A2'
    
    # 保存
    excel_path = os.path.join(output_dir, f"{base_name}_发票信息汇总表.xlsx")
    wb.save(excel_path)
    return excel_path

def process_batch(pdf_files, output_dir=None):
    """批量处理 PDF 文件"""
    print(f"=" * 60)
    print(f"发票 OCR 批量处理工具 v1.0")
    print(f"=" * 60)
    print(f"待处理文件：{len(pdf_files)} 个")
    
    results = []
    for i, pdf_path in enumerate(pdf_files, 1):
        print(f"\n[{i}/{len(pdf_files)}]")
        result = process_pdf(pdf_path, output_dir)
        if result:
            results.append(result)
    
    # 汇总统计
    if results:
        print(f"\n" + "=" * 60)
        print(f"批量处理完成")
        print(f"=" * 60)
        print(f"成功：{len(results)}/{len(pdf_files)}")
        
        total_invoices = sum(r["stats"]["total"] for r in results)
        total_amount = sum(r["stats"]["total_amount"] for r in results)
        print(f"总发票数：{total_invoices} 张")
        print(f"总金额：¥{total_amount:,.2f}")

def main():
    """主函数"""
    print("=" * 60)
    print("发票 OCR 识别工具 v1.0")
    print("=" * 60)
    
    # 检查依赖
    try:
        import fitz
        import pytesseract
        from PIL import Image
    except ImportError as e:
        print(f"错误：缺少依赖库 - {e}")
        print("请运行：pip3 install pymupdf pytesseract pillow openpyxl")
        sys.exit(1)
    
    # 命令行参数
    if len(sys.argv) < 2:
        print("\n使用方法:")
        print("  python3 invoice_ocr.py <PDF 文件路径> [输出目录]")
        print("  python3 invoice_ocr.py <PDF 文件 1> <PDF 文件 2> ... [输出目录]")
        print("\n示例:")
        print("  python3 invoice_ocr.py /home/ubuntu/凭证.pdf")
        print("  python3 invoice_ocr.py *.pdf /home/ubuntu/output")
        sys.exit(0)
    
    # 解析参数
    pdf_files = []
    output_dir = None
    
    for arg in sys.argv[1:]:
        if os.path.isdir(arg):
            output_dir = arg
        elif arg.endswith('.pdf'):
            pdf_files.extend(Path('.').glob(arg))
        else:
            if os.path.exists(arg):
                pdf_files.append(arg)
    
    if not pdf_files:
        print("错误：未找到 PDF 文件")
        sys.exit(1)
    
    # 处理
    if len(pdf_files) == 1:
        process_pdf(pdf_files[0], output_dir)
    else:
        process_batch(pdf_files, output_dir)
    
    print("\n完成！")

if __name__ == "__main__":
    main()
